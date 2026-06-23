import openpyxl
import csv
import random

INPUT = "/Users/rory.wickham/claude-projects/goodyearWork/goodyear_tires.xlsx"
OUTPUT = "/Users/rory.wickham/claude-projects/goodyearWork/goodyear_tires_feed.csv"

wb = openpyxl.load_workbook(INPUT)
ws = wb.active

headers = [cell.value for cell in ws[1]]
rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    rows.append(dict(zip(headers, row)))

out_headers = [
    "ID", "ItemGroupID", "Title", "Description",
    "Link", "imageURL", "AlternateImageURL",
    "ProductCategory", "Subcategory1", "Subcategory2", "Subcategory3",
    "ProductClass", "Manufacturer", "Color", "ColorSwatchURL",
    "Size", "TaxCode", "SalePrice", "ListPrice", "OnlineInventory",
    "Currency", "GTIN",
    "Rim Diameter", "Price Range", "Warranties & Guarantees",
]

def first_alt_image(thumb_images):
    if not thumb_images:
        return ""
    parts = str(thumb_images).split(" | ")
    return parts[1] if len(parts) > 1 else ""

def clean_price(price_str):
    if not price_str:
        return ""
    return str(price_str).replace("$", "").strip()

def strip_commas(value):
    if not value:
        return ""
    return str(value).replace(",", "")

out_rows = []
for r in rows:
    category = r.get("Categories", "") or ""
    parts = [p.strip() for p in category.split(">")]

    price = clean_price(r.get("Price", ""))

    out_rows.append({
        "ID": r.get("Variant ID", ""),
        "ItemGroupID": r.get("Master Product ID", ""),
        "Title": strip_commas(r.get("Name", "")),
        "Description": strip_commas(r.get("Tire Features", "")),
        "Link": r.get("Product URL", ""),
        "imageURL": r.get("Listing Image URL", ""),
        "AlternateImageURL": first_alt_image(r.get("All Image URLs", "")),
        "ProductCategory": category,
        "Subcategory1": parts[0] if len(parts) > 0 else "",
        "Subcategory2": parts[1] if len(parts) > 1 else "",
        "Subcategory3": parts[2] if len(parts) > 2 else "",
        "ProductClass": "Merchandise",
        "Manufacturer": r.get("Brand", ""),
        "Color": "Black",
        "ColorSwatchURL": "",
        "Size": r.get("Tire Size", ""),
        "TaxCode": "CLT",
        "SalePrice": price,
        "ListPrice": price,
        "OnlineInventory": "1000",
        "Currency": "USD",
        "GTIN": str(random.randint(1000000, 9999999)),
        "Rim Diameter": strip_commas(r.get("Rim Diameter", "")),
        "Price Range": strip_commas(r.get("Price Range", "")),
        "Warranties & Guarantees": strip_commas(r.get("Warranties & Guarantees", "")),
    })

out_rows = [r for r in out_rows if r["SalePrice"] and r["ListPrice"]]

# Strip commas from all string fields to prevent CSV column count mismatches
skip_fields = {"SalePrice", "ListPrice", "GTIN"}
for r in out_rows:
    for key, val in r.items():
        if key not in skip_fields and isinstance(val, str):
            r[key] = val.replace(",", "").replace("\n", " ").replace("\r", "")

with open(OUTPUT, "w", newline="\n", encoding="utf-8") as f:
    writer = csv.DictWriter(f, fieldnames=out_headers)
    writer.writeheader()
    writer.writerows(out_rows)

print(f"Written {len(out_rows)} rows to {OUTPUT}")
